#!/usr/bin/env python3
import argparse
import os
import re
import csv
from datetime import datetime, date, time
from openpyxl import load_workbook


def safe_name(name):
    name = name or "sheet"
    name = re.sub(r'[\\/*?:"<>|]', "_", name)
    name = name.strip()[:200]
    return name or "sheet"


def to_str(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def xlsx_to_csv(input_path, out_dir):
    wb = load_workbook(input_path, read_only=True, data_only=True)
    os.makedirs(out_dir, exist_ok=True)
    created = []
    used_names = {}
    for idx, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        base = f"{idx:02d}_{safe_name(sheet_name)}"
        # avoid collisions
        count = used_names.get(base, 0)
        used_names[base] = count + 1
        filename = base if count == 0 else f"{base}_{count}"
        csv_path = os.path.join(out_dir, f"{filename}.csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([to_str(v) for v in row])
        created.append(csv_path)
        print("Wrote:", csv_path)
    return created


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Convert .xlsx sheets to CSV files")
    p.add_argument("xlsx", help="Path to XLSX file")
    p.add_argument("--outdir", "-o", default="xlsx_csv_output", help="Output directory")
    args = p.parse_args()
    files = xlsx_to_csv(args.xlsx, args.outdir)
    print("Done. Created", len(files), "csv files in", args.outdir)
